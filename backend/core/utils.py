import io
import os
import pandas as pd
import numpy as np
from typing import Dict, Any, List, Optional

def get_dataframe_metadata(df: pd.DataFrame) -> Dict[str, Any]:
    """Extracts detailed metadata from a DataFrame for LLM context."""
    # Basic info
    info = {
        "shape": df.shape,
        "row_count": len(df),
        "column_count": len(df.columns),
        "columns": list(df.columns),
        "dtypes": {str(k): str(v) for k, v in df.dtypes.items()},
    }
    
    # Missing values summary
    null_counts = df.isnull().sum()
    info["missing_values"] = {
        col: int(count) for col, count in null_counts.items() if count > 0
    }
    
    # Detect date columns
    date_cols = []
    for col in df.columns:
        if df[col].dtype == 'object':
            try:
                sample = df[col].dropna().head(10)
                if not sample.empty:
                    pd.to_datetime(sample)
                    date_cols.append(col)
            except (ValueError, TypeError):
                continue
    info["potential_date_columns"] = date_cols
    
    # Categorical vs Numeric
    numeric_df = df.select_dtypes(include=[np.number])
    categorical_df = df.select_dtypes(include=['object', 'category'])
    
    info["numeric_columns"] = list(numeric_df.columns)
    info["categorical_columns"] = list(categorical_df.columns)
    
    # Categorical unique values (top 5)
    info["categorical_unique_counts"] = {
        col: {
            "unique_count": df[col].nunique(),
            "top_values": [str(v) if isinstance(v, (pd.Timestamp, np.datetime64)) else v for v in df[col].value_counts().head(5).index]
        }
        for col in categorical_df.columns
    }
    
    # Numeric summary statistics
    if not numeric_df.empty:
        stats = numeric_df.describe().to_dict()
        info["numeric_stats"] = {
            col: {k: float(v) if not np.isnan(v) and not isinstance(v, (pd.Timestamp, np.datetime64)) else (str(v) if isinstance(v, (pd.Timestamp, np.datetime64)) else None) for k, v in col_stats.items()}
            for col, col_stats in stats.items()
        }
    
    # Data preview (head 5) - ensure all values are serializable
    preview = df.head(5).to_dict(orient='records')
    for row in preview:
        for key, val in row.items():
            if isinstance(val, (pd.Timestamp, np.datetime64)):
                row[key] = str(val)
            elif pd.isna(val):
                row[key] = None
    info["preview"] = preview
    
    return info

def create_excel_report(df: pd.DataFrame, analysis_results: Dict[str, Any], sheet_name: str = "Data") -> bytes:
    """Creates a professional multi-sheet Excel report using XlsxWriter and returns bytes."""
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        # Write main data
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # Access workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]
        
        # Apply premium Navy style to table headers
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#1E3A8A',
            'font_color': 'white',
            'border': 1
        })
        for col_num, value in enumerate(df.columns.values):
            worksheet.write(0, col_num, value, header_format)
            
        # Adjust column widths automatically
        for i, col in enumerate(df.columns):
            max_len = max(df[col].astype(str).map(len).max(), len(str(col))) + 3
            max_len = min(max_len, 50)  # Caps width
            worksheet.set_column(i, i, max_len)
        
        # Write Analysis sheet
        analysis_sheet = workbook.add_worksheet("Analysis")
        title_format = workbook.add_format({'bold': True, 'font_size': 14, 'font_color': '#1E3A8A'})
        section_format = workbook.add_format({'bold': True, 'font_size': 12, 'bg_color': '#F3F4F6'})
        text_format = workbook.add_format({'text_wrap': True})
        
        analysis_sheet.write(0, 0, "Exodus Analysis Report", title_format)
        
        analysis_sheet.write(2, 0, "Executive Summary", section_format)
        analysis_sheet.write(3, 0, analysis_results.get("summary", "No summary generated."), text_format)
        
        if "findings" in analysis_results and isinstance(analysis_results["findings"], list):
            analysis_sheet.write(5, 0, "Key Findings", section_format)
            for idx, finding in enumerate(analysis_results["findings"]):
                analysis_sheet.write(6 + idx, 0, f"• {finding}", text_format)
                
        # Set analysis sheet column widths
        analysis_sheet.set_column(0, 0, 80)
        
    return output.getvalue()

def embed_chart_in_excel(report_bytes: bytes, charts: list, df: pd.DataFrame) -> bytes:
    """Converts Plotly figures to PNGs and embeds them into the Excel byte stream."""
    import openpyxl
    from openpyxl.drawing.image import Image
    
    in_file = io.BytesIO(report_bytes)
    wb = openpyxl.load_workbook(in_file)
    
    # Create or clean 'Charts' sheet
    if "Charts" not in wb.sheetnames:
        ws = wb.create_sheet("Charts")
    else:
        ws = wb["Charts"]
        
    ws.views.sheetView[0].showGridLines = True
    
    row_offset = 2
    for fig in charts:
        if fig is None:
            continue
        try:
            # Render Plotly fig to PNG bytes via Kaleido
            img_bytes = fig.to_image(format="png", width=700, height=450)
            img_io = io.BytesIO(img_bytes)
            img = Image(img_io)
            
            cell_loc = f"B{row_offset}"
            ws.add_image(img, cell_loc)
            row_offset += 24  # Space out charts
        except Exception as e:
            ws.cell(row=row_offset, column=2, value=f"Failed to render chart: {e}")
            row_offset += 2
            
    out_io = io.BytesIO()
    wb.save(out_io)
    return out_io.getvalue()

def find_matching_column(df: pd.DataFrame, values: Any) -> Optional[int]:
    if values is None:
        return None
    val_list = list(values)
    for col_idx, col_name in enumerate(df.columns):
        col_list = list(df[col_name])
        if len(col_list) == len(val_list):
            match = True
            for v1, v2 in zip(col_list, val_list):
                if pd.isna(v1) and pd.isna(v2):
                    continue
                try:
                    if hasattr(v1, 'item'):
                        v1 = v1.item()
                    if hasattr(v2, 'item'):
                        v2 = v2.item()
                    if isinstance(v1, float) and isinstance(v2, float):
                        if not np.isclose(v1, v2, equal_nan=True):
                            match = False
                            break
                    elif v1 != v2:
                        match = False
                        break
                except Exception:
                    match = False
                    break
            if match:
                return col_idx + 1
    return None

def extract_plotly_data(fig_val) -> pd.DataFrame:
    if fig_val is None or not hasattr(fig_val, 'data') or not fig_val.data:
        return pd.DataFrame()
        
    traces = fig_val.data
    x_label = "X"
    y_label = "Y"
    if hasattr(fig_val, 'layout'):
        if fig_val.layout.xaxis and fig_val.layout.xaxis.title and fig_val.layout.xaxis.title.text:
            x_label = fig_val.layout.xaxis.title.text
        if fig_val.layout.yaxis and fig_val.layout.yaxis.title and fig_val.layout.yaxis.title.text:
            y_label = fig_val.layout.yaxis.title.text

    dfs = []
    for idx, trace in enumerate(traces):
        x_data = getattr(trace, 'x', None)
        y_data = getattr(trace, 'y', None)
        if x_data is None or y_data is None:
            continue
        
        x_list = list(x_data)
        y_list = list(y_data)
        
        trace_name = getattr(trace, 'name', None) or f"{y_label} (Series {idx+1})"
        
        x_list = [v.item() if hasattr(v, 'item') else v for v in x_list]
        y_list = [v.item() if hasattr(v, 'item') else v for v in y_list]
        
        df_trace = pd.DataFrame({x_label: x_list, trace_name: y_list})
        dfs.append(df_trace)
        
    if not dfs:
        return pd.DataFrame()
        
    if len(dfs) == 1:
        return dfs[0]
        
    df_merged = dfs[0]
    for df_t in dfs[1:]:
        df_merged = pd.merge(df_merged, df_t, on=x_label, how='outer')
        
    return df_merged

def plotly_to_excel_chart(ws, fig_val, data_end_row: int, x_col_idx: int, y_col_indices: List[int]):
    from openpyxl.chart import BarChart, LineChart, ScatterChart, Reference, Series
    
    first_trace = fig_val.data[0]
    trace_type = getattr(first_trace, 'type', 'scatter')
    trace_mode = getattr(first_trace, 'mode', 'lines') if hasattr(first_trace, 'mode') else 'lines'
    
    if trace_type == 'bar':
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.varyColors = True
    elif trace_type in ('scatter', 'scattergl'):
        if 'lines' in trace_mode:
            chart = LineChart()
            chart.style = 13
        else:
            chart = ScatterChart()
            chart.style = 13
        chart.varyColors = True
    else:
        raise NotImplementedError(f"Unsupported native chart type: {trace_type}")
        
    title_text = "Chart"
    if hasattr(fig_val, 'layout') and fig_val.layout.title and fig_val.layout.title.text:
        title_text = fig_val.layout.title.text
    chart.title = title_text
    
    x_title = None
    y_title = None
    if hasattr(fig_val, 'layout'):
        if fig_val.layout.xaxis and fig_val.layout.xaxis.title and fig_val.layout.xaxis.title.text:
            x_title = fig_val.layout.xaxis.title.text
        if fig_val.layout.yaxis and fig_val.layout.yaxis.title and fig_val.layout.yaxis.title.text:
            y_title = fig_val.layout.yaxis.title.text
            
    if x_title:
        chart.x_axis.title = x_title
    if y_title:
        chart.y_axis.title = y_title
        
    cats = Reference(ws, min_col=x_col_idx, min_row=2, max_row=data_end_row)
    
    if isinstance(chart, (BarChart, LineChart)):
        for y_col in y_col_indices:
            data = Reference(ws, min_col=y_col, min_row=1, max_row=data_end_row)
            chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
    else:
        for y_col in y_col_indices:
            xvalues = Reference(ws, min_col=x_col_idx, min_row=2, max_row=data_end_row)
            yvalues = Reference(ws, min_col=y_col, min_row=2, max_row=data_end_row)
            series = Series(yvalues, xvalues, title_from_data=True)
            chart.series.append(series)
            
    cell_loc = f"A{data_end_row + 3}"
    ws.add_chart(chart, cell_loc)

def append_insight_to_report(report_path: str, result_val: Any, fig_val: Any = None):
    """Appends a new sheet with result_val and fig_val (native chart / PNG fallback) to the report_path excel file."""
    import openpyxl
    import base64
    from openpyxl.drawing.image import Image
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    if not os.path.exists(report_path):
        return
        
    wb = openpyxl.load_workbook(report_path)
    
    insight_count = sum(1 for name in wb.sheetnames if name.startswith("Insight_"))
    sheet_name = f"Insight_{insight_count + 1}"
    ws = wb.create_sheet(sheet_name)
    
    ws.views.sheetView[0].showGridLines = True
    
    fig_base64 = None
    if fig_val is not None:
        try:
            import base64
            img_bytes = fig_val.to_image(format="png", width=700, height=450)
            fig_base64 = base64.b64encode(img_bytes).decode('utf-8')
        except Exception as e:
            print(f"Failed to encode fig to base64 in utils: {e}")
            
    data_written = False
    data_end_row = 1
    x_col_idx = None
    y_col_indices = []
    
    if isinstance(result_val, (pd.DataFrame, pd.Series)):
        df_to_write = result_val.to_frame() if isinstance(result_val, pd.Series) else result_val
        for r_idx, row in enumerate(dataframe_to_rows(df_to_write, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                if pd.isna(value):
                    val = ""
                elif isinstance(value, (int, float, np.integer, np.floating)):
                    val = value
                else:
                    val = str(value)
                ws.cell(row=r_idx, column=c_idx, value=val)
        data_end_row = df_to_write.shape[0] + 1
        data_written = True
        
        if fig_val is not None and hasattr(fig_val, 'data') and fig_val.data:
            try:
                first_trace = fig_val.data[0]
                x_col_idx = find_matching_column(df_to_write, getattr(first_trace, 'x', None))
                for trace in fig_val.data:
                    y_idx = find_matching_column(df_to_write, getattr(trace, 'y', None))
                    if y_idx is not None:
                        y_col_indices.append(y_idx)
            except Exception as e:
                print(f"Error matching columns: {e}")
                x_col_idx = None
                y_col_indices = []

    if fig_val is not None and (not data_written or x_col_idx is None or not y_col_indices):
        ws.delete_rows(1, ws.max_row + 1)
        try:
            df_chart = extract_plotly_data(fig_val)
            if not df_chart.empty:
                for r_idx, row in enumerate(dataframe_to_rows(df_chart, index=False, header=True), 1):
                    for c_idx, value in enumerate(row, 1):
                        if pd.isna(value):
                            val = ""
                        elif isinstance(value, (int, float, np.integer, np.floating)):
                            val = value
                        else:
                            val = str(value)
                        ws.cell(row=r_idx, column=c_idx, value=val)
                data_end_row = df_chart.shape[0] + 1
                data_written = True
                x_col_idx = 1
                y_col_indices = list(range(2, df_chart.shape[1] + 1))
        except Exception as e:
            print(f"Failed to extract and write chart data: {e}")
            
    if not data_written and result_val is not None:
        ws.cell(row=1, column=1, value=str(result_val))
        data_end_row = 1
        data_written = True

    chart_added = False
    if fig_val is not None and x_col_idx is not None and y_col_indices:
        try:
            plotly_to_excel_chart(ws, fig_val, data_end_row, x_col_idx, y_col_indices)
            chart_added = True
        except Exception as e:
            print(f"Failed to generate native chart: {e}. Falling back to PNG.")
            
    if fig_val is not None and not chart_added and fig_base64:
        try:
            img_bytes = base64.b64decode(fig_base64)
            img_io = io.BytesIO(img_bytes)
            img = Image(img_io)
            cell_loc = f"A{data_end_row + 3}"
            ws.add_image(img, cell_loc)
        except Exception as e:
            ws.cell(row=data_end_row + 3, column=1, value=f"Failed to embed chart: {e}")
            
    wb.save(report_path)
